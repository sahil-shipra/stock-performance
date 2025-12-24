import os
import pandas as pd


def import_csv(file_path):
    """
    Import trading data from Excel (.xlsx) or CSV (.csv) file
    Returns a tuple: (main_df, summary_df, equity_curve_df) where summary_df and equity_curve_df are None if not found
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == '.xlsx':
        # Read the main data (first sheet or default)
        df = pd.read_excel(file_path)
        print(f"✓ Successfully imported Excel file: {file_path}")

        # Try to read the Summary and Equity Curve worksheets if they exist
        summary_df = None
        equity_curve_df = None
        try:
            import openpyxl
            # Use openpyxl to read values (not formulas) from Summary sheet
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Read Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                # Convert worksheet to list of lists
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

                # Create DataFrame from the data
                if data:
                    # First row as headers
                    summary_df = pd.DataFrame(data[1:], columns=data[0])
                    print(f"✓ Found and preserved 'Summary' worksheet with values")

            # Read Equity Curve sheet
            if 'Equity Curve' in wb.sheetnames:
                ws = wb['Equity Curve']
                # Convert worksheet to list of lists
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

                # Create DataFrame from the data
                if data:
                    # First row as headers
                    equity_curve_df = pd.DataFrame(data[1:], columns=data[0])
                    print(f"✓ Found and preserved 'Equity Curve' worksheet with values")

            wb.close()
        except Exception as e:
            print(f"⚠ Could not read Summary/Equity Curve worksheet: {e}")

        return df, summary_df, equity_curve_df
    elif file_extension == '.csv':
        df = pd.read_csv(file_path)
        print(f"✓ Successfully imported CSV file: {file_path}")
        return df, None, None
    else:
        raise ValueError(
            f"Unsupported file format: {file_extension}. Please use .xlsx or .csv")

