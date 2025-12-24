from io import BytesIO
import pandas as pd


def read_upload_excel(contents: BytesIO):

    df = pd.read_excel(contents)
    summary_df = None
    equity_curve_df = None

    try:
        import openpyxl
        # Use openpyxl to read values (not formulas) from Summary sheet
        wb = openpyxl.load_workbook(contents, data_only=True)

        # Read Summary sheet
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            # Convert worksheet to list of lists
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            # Create DataFrame from the data
            if data:
                # First row as headers
                summary_df = pd.DataFrame(data[1:], columns=data[0])
                print(f"✓ Found and preserved 'Summary' worksheet with values")

        # Read Equity Curve sheet
        if 'Equity Curve' in wb.sheetnames:
            ws = wb['Equity Curve']
            # Convert worksheet to list of lists
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            # Create DataFrame from the data
            if data:
                # First row as headers
                equity_curve_df = pd.DataFrame(data[1:], columns=data[0])
                print(f"✓ Found and preserved 'Equity Curve' worksheet with values")

        wb.close()
    except Exception as e:
        print(f"⚠ Could not read Summary/Equity Curve worksheet: {e}")

    return df, summary_df, equity_curve_df
